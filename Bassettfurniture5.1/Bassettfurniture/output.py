
def out(fname,diction,csv=False,json=False,xl=False):
    import os
    if type(diction) == dict:
        if csv:
            file = fname + '.csv'
            if not file in os.listdir():
                fp = open(file,'x')
                fp.close()
                with open(file,'w',encoding='utf-8') as f:
                    f.write(','.join([f'"{k}"' if ',' in k else k for k in diction.keys()])+'\n')

            with open(file,'a',encoding='utf-8') as f:
                f.write(','.join([f'"{v}"' if ',' in v else v for v in diction.values()])+'\n')
        if json:
            import json
            file = fname + '.json'
            if not file in os.listdir():
                fp = open(file, 'x')
                fp.close()

            with open(file, 'r', encoding='utf-8') as f:
                data = f.read()
            with open(file, 'w', encoding='utf-8') as f:
                if data:
                    data = json.loads(data)
                    new_data = [*data, diction]
                    f.write(json.dumps(new_data, indent=2))
                else:
                    f.write(json.dumps([diction], indent=2))
        if xl:
            from openpyxl import Workbook, load_workbook
            file = fname + '.xlsx'
            if not file in os.listdir():
                wb = Workbook(file)
                wb.create_sheet()
                sheet1 = wb.active
                sheet1.append(list(diction.keys()))
                wb.save(file)

            wb = load_workbook(file)
            sheet1 = wb.active
            sheet1.append(list(diction.values()))
            wb.save(file)
def outlist(fname,dictionlist,csv=False,json=False,xl=False):
    import os
    if type(dictionlist) == list:
        if csv:
            file = fname + '.csv'
            if not file in os.listdir():
                fp = open(file,'x')
                fp.close()
                with open(file,'w',encoding='utf-8') as f:
                    f.write(','.join([f'"{k}"' if ',' in k else k for k in dictionlist[0].keys()])+'\n')

            with open(file,'a',encoding='utf-8') as f:
                for d in dictionlist:
                    f.write(','.join([f'"{v}"' if ',' in v else v for v in d.values()])+'\n')
        if json:
            import json
            file = fname + '.json'

            if not file in os.listdir():
                fp = open(file, 'x')
                fp.close()

            with open(file, 'r', encoding='utf-8') as f:
                data = f.read()
            with open(file, 'w', encoding='utf-8') as f:
                if data:
                    data = json.loads(data)
                    new_data = [*data, *dictionlist]
                    f.write(json.dumps(new_data, indent=2))
                else:
                    f.write(json.dumps(dictionlist, indent=2))
        if xl:
            from openpyxl import Workbook, load_workbook
            file = fname + '.xlsx'

            if not file in os.listdir():
                wb = Workbook(file)
                wb.create_sheet()
                sheet1 = wb.active
                sheet1.append(list(dictionlist[0].keys()))
                wb.save(file)

            wb = load_workbook(file)
            sheet1 = wb.active
            for d in dictionlist:
                sheet1.append(list(d.values()))
            wb.save(file)

if __name__ == '__main__':
    pass